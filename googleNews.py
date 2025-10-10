from functools import partial

import pytz
import openpyxl
import requests
import pandas as pd
from common import *
import concurrent.futures
from openpyxl import Workbook
from datetime import timedelta
from openpyxl.utils.dataframe import dataframe_to_rows
from dateutil import parser


# Define a function to process each row
def process_row_with_custom_news_channel(row_value):
    name_value = row_value[2].value
    if name_value is None:
        return None

    # Add the value to the list
    name_value = name_value.replace('&', 'and')

    url_value = "https://news.google.com/rss/search?q='" + name_value + "'&hl=en-IN&gl=IN&ceid=IN:en"
    response_value = requests.get(url_value)
    return response_value.text


def defined_keyword_base_search(title):
    signal_found = "neutral"
    keyword_found = ""
    if 'profile' in title:
        signal_found = "positive"
        keyword_found += "profile,"
    if 'deal' in title:
        signal_found = "positive"
        keyword_found += "deal,"
    if 'approve' in title:
        signal_found = "positive"
        keyword_found += "approve,"
    if 'gain' in title:
        signal_found = "positive"
        keyword_found += "gain,"
    if 'rise' in title:
        signal_found = "positive"
        keyword_found += "rise,"
    if 'win' in title:
        signal_found = "positive"
        keyword_found += "win,"
    if 'partnership' in title:
        signal_found = "positive"
        keyword_found += "partnership,"
    if 'partner' in title:
        signal_found = "positive"
        keyword_found += "partner,"
    if 'order' in title:
        signal_found = "positive"
        keyword_found += "order,"
    if 'soar' in title:
        signal_found = "positive"
        keyword_found += "soar,"
    if 'climb' in title:
        signal_found = "positive"
        keyword_found += "climb,"
    if 'ascend' in title:
        signal_found = "positive"
        keyword_found += "ascend,"
    if 'lift' in title:
        signal_found = "positive"
        keyword_found += "lift,"
    if 'authorize' in title:
        signal_found = "positive"
        keyword_found += "authorize,"
    if 'empower' in title:
        signal_found = "positive"
        keyword_found += "empower,"
    if 'increase' in title:
        signal_found = "positive"
        keyword_found += "increase,"
    if 'mount' in title:
        signal_found = "positive"
        keyword_found += "lift,"
    if 'loss' in title:
        signal_found = "negative"
        keyword_found += "loss,"
    if 'reject' in title:
        signal_found = "negative"
        keyword_found += "reject,"
    if 'down' in title:
        signal_found = "negative"
        keyword_found += "down,"
    if 'fall' in title:
        signal_found = "negative"
        keyword_found += "fall,"
    if 'low' in title:
        signal_found = "negative"
        keyword_found += "low,"
    return signal_found, keyword_found


def is_within_n_time_units_ist_google(date_str, value, unit):
    try:
        # Parse the given date string
        given_date = parser.parse(date_str)

        # Get the current date and time
        current_date = datetime.now()

        # Calculate the difference between the current date and the given date
        difference = current_date - given_date

        # Determine the timedelta based on the unit
        if unit == 'hours':
            time_delta = timedelta(hours=value)
        elif unit == 'minutes':
            time_delta = timedelta(minutes=value)
        elif unit == 'seconds':
            time_delta = timedelta(seconds=value)
        else:
            raise ValueError("Invalid unit. Use 'hours', 'minutes', or 'seconds'.")

        # Check if the difference is within the specified time unit
        return difference <= time_delta
    except ValueError as e:
        print(f"Error: {e}")
        return False


def parse_response(current,time_cycle_for_report, time_unit_for_report):
    valid_xml_flag, root = is_valid_xml(current)
    if not valid_xml_flag:
        return None

    channel_title = root.find('channel/title').text.strip() if root.find('channel/title') is not None else None
    if channel_title:
        channel_title = channel_title.replace('\'\" - Google News', '').replace('\"\'', '')

    items_data = []
    for item in root.findall('.//item'):
        title = item.find('title').text.strip() if item.find('title') is not None else None
        pub_date = item.find('pubDate').text.strip() if item.find('pubDate') is not None else None
        # UTC timestamp string
        utc_timestamp_str = pub_date

        # Define UTC and IST timezones
        utc_timezone = pytz.timezone('UTC')
        ist_timezone = pytz.timezone('Asia/Kolkata')  # Indian Standard Time

        # Convert UTC string to datetime object
        utc_datetime = datetime.strptime(utc_timestamp_str, '%a, %d %b %Y %H:%M:%S %Z')

        # Localize UTC datetime to UTC timezone
        utc_datetime = utc_timezone.localize(utc_datetime)

        # Convert localized UTC datetime to IST datetime
        ist_datetime = utc_datetime.astimezone(ist_timezone)

        # Format IST datetime as string
        ist_timestamp_str = ist_datetime.strftime('%a, %d %b %Y %H:%M:%S %Z')

        if not is_within_n_time_units_ist_google(ist_timestamp_str, time_cycle_for_report, time_unit_for_report):
            continue

        source_elem = item.find('source')
        source_name = source_elem.text.strip() if source_elem is not None else None
        source_url = source_elem.get('url') if source_elem is not None else None

        actual_title = title
        title = title.replace('NDTV Profit', '').lower()

        chat_content = "Response in \"positive\" or \"negative\" or \"neutral\" for this statement:\"" + title + "\""
        chat_completion = open_client.chat.completions.create(
            messages=[{"role": "system",
                       "content": chat_content}],
            model="gpt-4o-mini",
        )

        chat_gpt_response = chat_completion.choices[0].message.content.lower()
        signal_found, keyword_found = defined_keyword_base_search(title)

        items_data.append({
            'channel_title': channel_title,
            'title': actual_title,
            'pubDate': ist_timestamp_str,
            'source_name': source_name,
            'source_url': source_url,
            'chat_gpt_response': chat_gpt_response,
            'signal_found': signal_found,
            'keyword_found': keyword_found
        })

    return items_data


# Function to find all matching items
def find_items(items_data, criteria):
    # Ensure criteria is a string
    if not isinstance(criteria, str):
        raise TypeError("Criteria must be a string")

    matching_items = []

    for item in items_data:
        # Ensure item is a dictionary
        if not isinstance(item, dict):
            continue

        # Check if criteria is in channel_title or description
        title_match = criteria.lower() in item.get('channel_title', '').lower()
        description_match = criteria.lower() in item.get('description', '').lower()

        if title_match or description_match:
            matching_items.append(item)

    return matching_items


def auto_adjust_column_widths(sheet_local):
    for column in sheet_local.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Get the column letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length
        sheet_local.column_dimensions[column_letter].width = adjusted_width


def apply_filters(sheet_local, columns):
    # Apply filter to the specified columns
    sheet_local.auto_filter.ref = sheet_local.dimensions  # Apply filter to the entire sheet_local
    sheet_local.auto_filter.add_filter_column(0, columns)  # Add filter to all columns


def sort_by_column(sheet_local, column_index, ascending=False):
    # Read data into DataFrame
    data_local = list(sheet_local.values)
    columns = data_local[0]  # Header row
    df = pd.DataFrame(data_local[1:], columns=columns)

    # Sort DataFrame by column
    df = df.sort_values(by=columns[column_index], ascending=ascending)

    # Clear the sheet and write sorted data back
    for row in sheet_local.iter_rows(min_row=1, max_row=sheet_local.max_row, max_col=sheet_local.max_column):
        for cell in row:
            cell.value = None

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            sheet_local.cell(row=r_idx, column=c_idx, value=value)


def execute_google_core(time_cycle_for_report, time_unit_for_report):
    # Define variable to load the dataframe
    # Open the spreadsheet
    workbook = openpyxl.load_workbook("./inputfile/MCAP28032024.xlsx")

    # Get the first sheet
    sheet = workbook.worksheets[0]

    # Create a list to store the values
    responses = []

    # Use ThreadPoolExecutor to run requests in parallel
    with concurrent.futures.ThreadPoolExecutor() as executor:
        # Submit each row for processing concurrently
        future_to_row = {executor.submit(process_row_with_custom_news_channel, row): row for row in sheet}

        # Iterate through completed futures to gather responses
        for future in concurrent.futures.as_completed(future_to_row):
            row = future_to_row[future]
            try:
                response = future.result()
                if response is not None:
                    responses.append(response)
            except Exception as exc:
                print(f"Exception occurred: {exc}")

    # Now 'responses' contains all the fetched responses
    final_excel_data = []

    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        parse_with_params = partial(parse_response, time_cycle_for_report=time_cycle_for_report,
                                    time_unit_for_report=time_unit_for_report)
        results = list(executor.map(parse_with_params, responses))

    # Flatten the list of lists into a single list
    for result in results:
        if result:
            final_excel_data.extend(result)

    # Write to Excel
    wb = Workbook()
    ws = wb.active

    # Write headers
    headers = ['Channel Title', 'Title', 'Published Date', 'Source',
               # 'Source URL',
               'chatGPT', 'Signal', 'Keyword Found']
    ws.append(headers)

    # Write data from responses
    for final_data in final_excel_data:
        data = [
            final_data['channel_title'],
            final_data['title'],
            final_data['pubDate'],
            final_data['source_name'],
            # final_data['source_url'],
            final_data['chat_gpt_response'],
            final_data['signal_found'],
            final_data['keyword_found']
        ]
        ws.append(data)

    # Save the workbook
    excel_file = 'rss_feed_data.xlsx'
    wb.save(excel_file)

    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook('rss_feed_data.xlsx')
    sheet = workbook.active

    # Apply filters and adjust column widths
    sort_by_column(sheet, column_index=2, ascending=False)  # Assuming 'pubDate' is the third column
    apply_filters(sheet, list(range(sheet.max_column)))
    auto_adjust_column_widths(sheet)

    # Save the workbook
    excel_file = generate_filename(prefix='share_market', extension='xlsx')

    workbook.save(excel_file)

    # Infinite loop to execute the function repeatedly
    # Replace with your own credentials and details
    api_id = '26456121'
    api_hash = '74df6d467ed43266c62111554fcb6e90'
    phone_number = '+919322504136'  # Your Telegram phone number
    target = '+919029085929'  # The recipient's username or phone number
    message = 'Hello, please find the attached file.'
    attachment_path = excel_file  # Path to your file

    send_telegram_message_with_attachment(api_id, api_hash, phone_number, target, message, attachment_path)
